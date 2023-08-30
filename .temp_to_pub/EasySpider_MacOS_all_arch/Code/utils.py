# 控制流程的暂停和继续

import csv
import datetime
import json
import os
import sys
import re
import time
import uuid
# import keyboard
from openpyxl import Workbook, load_workbook
import requests
from urllib.parse import urlparse
import pymysql
from lxml import etree

def is_valid_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except ValueError:
        return False

def lowercase_tags_in_xpath(xpath):
    return re.sub(r"([A-Z]+)(?=[\[\]//]|$)", lambda x: x.group(0).lower(), xpath)


def on_press_creator(press_time, event):
    def on_press(key):
        try:
            if key.char == 'p':
                if press_time["is_pressed"] == False: # 没按下p键时，记录按下p键的时间
                    press_time["duration"] = time.time()
                    press_time["is_pressed"] = True
                else: # 按下p键时，判断按下p键的时间是否超过2.5秒
                    duration = time.time() - press_time["duration"]
                    if duration > 2:
                        if event._flag == False:
                            print("任务执行中，长按p键暂停执行。")
                            print("Task is running, long press 'p' to pause.")
                            # 设置Event的值为True，使得线程b可以继续执行
                            event.set()
                        else:
                            # 设置Event的值为False，使得线程b暂停执行
                            print("任务已暂停，长按p键继续执行...")
                            print("Task paused, long press 'p' to continue...")
                            event.clear()
                        press_time["duration"] = time.time()
                        press_time["is_pressed"] = False
                    # print("按下p键时间：", press_time["duration"])
        except:
            pass
    return on_press

def on_release_creator(event, press_time):
    def on_release(key):
        try:
            # duration = time.time() - press_time["duration"]
            # # print("松开p键时间：", time.time(), "Duration: ", duration)
            # if duration > 2.5 and key.char == 'p':
            #     if event._flag == False:
            #         print("任务执行中，按p键暂停执行。")
            #         print("Task is running, press 'p' to pause.")
            #         # 设置Event的值为True，使得线程b可以继续执行
            #         event.set()
            #     else:
            #         # 设置Event的值为False，使得线程b暂停执行
            #         print("任务已暂停，按p键继续执行...")
            #         print("Task paused, press 'p' to continue...")
            #         event.clear()
            #     press_time["duration"] = time.time()
            press_time["is_pressed"] = False
        except:
            pass
    return on_release


# def check_pause(key, event):
#     while True:
#         if keyboard.is_pressed(key):  # 按下p键，暂停程序
#             if event._flag == False:
#                 print("任务执行中，长按p键暂停执行。")
#                 print("Task is running, long press 'p' to pause.")
#                 # 设置Event的值为True，使得线程b可以继续执行
#                 event.set()
#             else:
#                 # 设置Event的值为False，使得线程b暂停执行
#                 print("任务已暂停，长按p键继续执行...")
#                 print("Task paused, press 'p' to continue...")
#                 event.clear()
#         time.sleep(1)  # 每秒检查一次


def download_image(url, save_directory):
    # 定义浏览器头信息
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    if is_valid_url(url):
        # 发送 GET 请求获取图片数据
        response = requests.get(url, headers=headers)

        # 检查响应状态码是否为成功状态
        if response.status_code == requests.codes.ok:
            # 提取文件名
            file_name = url.split('/')[-1].split("?")[0]

            # 生成唯一的新文件名
            new_file_name = file_name + '_' + \
                str(uuid.uuid4()) + '_' + file_name

            # 构建保存路径
            save_path = os.path.join(save_directory, new_file_name)

            # 保存图片到本地
            with open(save_path, 'wb') as file:
                file.write(response.content)

            print("图片已成功下载到:", save_path)
            print("The image has been successfully downloaded to:", save_path)
        else:
            print("下载图片失败，请检查此图片链接是否有效:", url)
            print(
                "Failed to download image, please check if this image link is valid:", url)
    else:
        print("下载图片失败，请检查此图片链接是否有效:", url)
        print("Failed to download image, please check if this image link is valid:", url)


def get_output_code(output):
    try:
        if output.find("rue") != -1:  # 如果返回值中包含true
            code = 1
        else:
            code = int(output)
    except:
        code = 0
    return code

# 判断字段是否为空


def isnull(s):
    return len(s) != 0

def new_line(outputParameters, maxViewLength, record):
    line = []
    i = 0
    for value in outputParameters.values():
        line.append(value)
        if record[i]:
            print(value[:maxViewLength], " ", end="")
        i += 1
    print("")
    return line

def write_to_csv(file_name, data, record):
    with open(file_name, 'a', encoding='utf-8-sig', newline="") as f:
        f_csv = csv.writer(f)
        for line in data:
            to_write = []
            for i in range(len(line)):
                if record[i]:
                    to_write.append(line[i])
            f_csv.writerow(to_write)
        f.close()


def write_to_excel(file_name, data, types, record):
    first = False
    if os.path.exists(file_name):
        # 加载现有的工作簿
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        # 创建新的工作簿和工作表
        wb = Workbook()
        ws = wb.active
        first = True
    # 追加数据到工作表
    for line in data:
        if not first: # 如果不是第一行，需要转换数据类型
            for i in range(len(line)):
                if types[i] == "int" or types[i] == "bigInt":
                    try:
                        line[i] = int(line[i])
                    except:
                        line[i] = 0
                elif types[i] == "double":
                    try:
                        line[i] = float(line[i])
                    except:
                        line[i] = 0.0
        else:
            first = False
        to_write = []
        for i in range(len(line)):
            if record[i]:
                to_write.append(line[i])
        ws.append(to_write)
    # 保存工作簿
    wb.save(file_name)





class Time:
    def __init__(self, type1=""):
        self.t = int(round(time.time() * 1000))
        self.type = type1

    def end(self):
        at = int(round(time.time() * 1000))
        print("Time used for", self.type, ":", at - self.t, "ms")


class myMySQL:
    def __init__(self, config_file="mysql_config.json"):
        # 读取配置文件
        try:
            if sys.platform == "darwin":
                if config_file.find("./") >= 0:
                    config_file = config_file.replace("./", "")
                config_file = os.path.expanduser("~/Library/Application Support/EasySpider/" + config_file)
            print("MySQL config file path: ", config_file)
            with open(config_file, 'r') as f:
                config = json.load(f)
                host = config["host"]
                port = config["port"]
                user = config["username"]
                passwd = config["password"]
                db = config["database"]
        except Exception as e:
            print("读取配置文件失败，请检查配置文件："+config_file+"是否存在。")
            print("Failed to read configuration file, please check if the configuration file: "+config_file+" exists.")
            print(e)
        try:
            self.conn = pymysql.connect(
            host=host, port=port, user=user, passwd=passwd, db=db)
            print("成功连接到数据库。")
            print("Successfully connected to the database.")
        except:
            print("连接数据库失败，请检查配置文件是否正确。")
            print("Failed to connect to the database, please check if the configuration file is correct.")
    
    def create_table(self, table_name, parameters):
        self.table_name = table_name
        self.field_sql = "("
        cursor = self.conn.cursor()
        # 检查表是否存在
        cursor.execute("SHOW TABLES LIKE '%s'" % table_name)
        result = cursor.fetchone()

        sql = "CREATE TABLE " + table_name + " (_id INT AUTO_INCREMENT PRIMARY KEY, "
        for item in parameters:
            if item["recordASField"]:
                name = item['name']
                if item['type'] == 'int':
                    sql += f"{name} INT, "
                elif item['type'] == 'double':
                    sql += f"{name} DOUBLE, "
                elif item['type'] == 'text':
                    sql += f"{name} TEXT, "
                elif item['type'] == 'mediumText':
                    sql += f"{name} MEDIUMTEXT, "
                elif item['type'] == 'longText':
                    sql += f"{name} LONGTEXT, "
                elif item['type'] == 'datetime':
                    sql += f"{name} DATETIME, "
                elif item['type'] == 'date':
                    sql += f"{name} DATE, "
                elif item['type'] == 'time':
                    sql += f"{name} TIME, "
                elif item['type'] == 'varchar':
                    sql += f"{name} VARCHAR(255), "
                elif item['type'] == 'bigInt':
                    sql += f"{name} BIGINT, "
                self.field_sql += f"{name}, "
        # 移除最后的逗号并添加闭合的括号
        sql = sql.rstrip(', ') + ")"
        self.field_sql = self.field_sql.rstrip(', ') + ")"

        # 如果表不存在，创建它
        if not result:
            # 执行SQL命令
            cursor.execute(sql)
        else:
            print("数据表" + table_name + "已存在。")
            print("The data table " + table_name + " already exists.")
        cursor.close()

    def write_to_mysql(self, OUTPUT, record, types):
        # 创建一个游标对象
        cursor = self.conn.cursor()

        for line in OUTPUT:
            for i in range(len(line)):
                if types[i] == "int" or types[i] == "bigInt":
                    try:
                        line[i] = int(line[i])
                    except:
                        line[i] = 0
                elif types[i] == "double":
                    try:
                        line[i] = float(line[i])
                    except:
                        line[i] = 0.0
                elif types[i] == "datetime":
                    try:
                        line[i] = datetime.datetime.strptime(line[i], '%Y-%m-%d %H:%M:%S')
                    except:
                        line[i] = datetime.datetime.strptime("1970-01-01 00:00:00", '%Y-%m-%d %H:%M:%S')
                elif types[i] == "date":
                    try:
                        line[i] = datetime.datetime.strptime(line[i], '%Y-%m-%d')
                    except:
                        line[i] = datetime.datetime.strptime("1970-01-01", '%Y-%m-%d')
                elif types[i] == "time":
                    try:
                        line[i] = datetime.datetime.strptime(line[i], '%H:%M:%S')
                    except:
                        line[i] = datetime.datetime.strptime("00:00:00", '%H:%M:%S')
            to_write = []
            for i in range(len(line)):
                if record[i]:
                    to_write.append(line[i])
            # 构造插入数据的 SQL 语句
            sql = f"INSERT INTO "+ self.table_name +" "+self.field_sql+" VALUES ("
            for item in to_write:
                sql += "%s, "
            # 移除最后的逗号并添加闭合的括号
            sql = sql.rstrip(', ') + ")"
            # 执行 SQL 语句
            try:
                cursor.execute(sql, to_write)
            except Exception as e:
                print("Error:", e)
                print("Error SQL:", sql, to_write)
                print("插入数据库错误，请查看以上的错误提示，然后检查数据的类型是否正确，是否文本过长（超过一万的文本类型要设置为大文本）。")
                print("Inserting database error, please check the above error, and then check whether the data type is correct, whether the text is too long (text type over 10,000 should be set to large text).")
                print("重新执行任务时，请删除数据库中的数据表" + self.table_name + "，然后再次运行程序。")
                print("When re-executing the task, please delete the data table " + self.table_name + " in the database, and then run the program again.")

        # 提交到数据库执行
        self.conn.commit()

        # 关闭游标和连接
        cursor.close()
    
    def close(self):
        self.conn.close()
        print("成功关闭数据库。")
        print("Successfully closed the database.")
